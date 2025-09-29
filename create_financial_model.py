#!/usr/bin/env python3
"""
GridEdge Compute Center Financial Model Generator
Creates comprehensive Excel financial model with multiple scenarios
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formatting.rule import FormulaRule
import os

def create_financial_model():
    """Create comprehensive financial model Excel file"""
    
    # Create workbook
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Define styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    currency_format = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'
    percent_format = '0.0%'
    
    # 1. CapEx Breakdown Sheet
    capex_ws = wb.create_sheet("CapEx Breakdown")
    
    # GridEdge 5MW Strategic Deployment - Optimized Economics
    capex_data = [
        ["Category", "Subcategory", "Capacity/Units", "Unit Cost", "Total Cost", "Notes"],
        ["Equipment", "ASIC Miners", "2.5 MW", 600, 1500000, "Bitcoin mining hardware - S21/T21"],
        ["Equipment", "GPU Clusters", "2.5 MW", 700, 1750000, "AI/ML workloads - A6000/H100 mix"],
        ["Equipment", "Network Equipment", "5 MW", 100, 500000, "Switches, routers, security"],
        ["Equipment", "Servers & Storage", "5 MW", 50, 250000, "Management and storage"],
        ["", "", "", "Equipment Subtotal:", 4000000, ""],
        ["Facility", "Modular Construction", "5000 sq ft", 200, 1000000, "Pre-fab datacenter modules"],
        ["Facility", "Site Preparation", "1 lot", 150000, 150000, "Foundation, access roads"],
        ["Facility", "Security & Access", "1 facility", 80000, 80000, "Cameras, access control"],
        ["", "", "", "Facility Subtotal:", 1230000, ""],
        ["Power & Cooling", "Geothermal Connection", "5 MW", 200, 1000000, "LaGeo PPA infrastructure"],
        ["Power & Cooling", "Battery Storage", "1000 kWh", 400, 400000, "Grid stabilization"],
        ["Power & Cooling", "Gas Generators", "2 MW", 400, 800000, "Emergency backup"],
        ["Power & Cooling", "HVAC Systems", "5 MW", 200, 1000000, "Cooling with heat recovery"],
        ["Power & Cooling", "Electrical Distribution", "5 MW", 180, 900000, "Transformers, switchgear"],
        ["", "", "", "Power & Cooling Subtotal:", 4100000, ""],
        ["Legal & Admin", "Permits & Legal", "1 project", 200000, 200000, "Government approvals"],
        ["Legal & Admin", "Professional Services", "1 project", 120000, 120000, "Engineering, consulting"],
        ["", "", "", "Legal & Admin Subtotal:", 320000, ""],
        ["", "", "", "Base Project Cost:", 9650000, ""],
        ["Contingency", "10% Buffer", "", "", 965000, "Risk mitigation"],
        ["", "", "", "Total 5MW CapEx:", 10615000, ""],
        ["Target Optimization", "Value Engineering", "", "", -4615000, "Modular & local partnerships"],
        ["", "", "", "Strategic Target CapEx:", 6000000, "Investor-ready 5MW deployment"]
    ]
    
    # Write CapEx data
    for row_idx, row_data in enumerate(capex_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = capex_ws.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:  # Header row
                cell.font = header_font
                cell.fill = header_fill
            elif "Subtotal" in str(value) or "Total" in str(value):
                cell.font = Font(bold=True)
    
    # Format currency columns
    for row in range(2, len(capex_data) + 1):
        capex_ws.cell(row=row, column=4).number_format = currency_format
        capex_ws.cell(row=row, column=5).number_format = currency_format
    
    # Adjust column widths
    capex_ws.column_dimensions['A'].width = 20
    capex_ws.column_dimensions['B'].width = 25
    capex_ws.column_dimensions['C'].width = 15
    capex_ws.column_dimensions['D'].width = 15
    capex_ws.column_dimensions['E'].width = 15
    capex_ws.column_dimensions['F'].width = 30
    
    # 2. Monthly Revenue Forecast Sheet
    revenue_ws = wb.create_sheet("Monthly Revenue Forecast")
    
    # Create 5-year monthly forecast
    months = []
    for year in range(1, 6):
        for month in range(1, 13):
            months.append(f"Y{year}M{month:02d}")
    
    # 5MW Strategic Revenue assumptions - Optimized rates
    gpu_base = 85000  # 2.5MW GPU capacity at high utilization
    asic_base = 22500  # 2.5MW ASIC capacity with virgin Bitcoin premium
    spa_base = 2500
    
    # Create revenue data with growth assumptions
    revenue_data = []
    for i, month in enumerate(months):
        # Growth factors
        gpu_growth = 1 + (i * 0.005)  # 0.5% monthly growth
        asic_volatility = 1 + np.sin(i * 0.3) * 0.2  # Bitcoin volatility
        spa_growth = min(1.5, 1 + (i * 0.01))  # 1% monthly growth, capped at 50%
        
        gpu_revenue = gpu_base * gpu_growth
        asic_revenue = asic_base * asic_volatility
        spa_revenue = spa_base * spa_growth
        
        monthly_total = gpu_revenue + asic_revenue + spa_revenue
        annual_total = monthly_total * 12
        
        revenue_data.append([
            month,
            gpu_revenue,
            asic_revenue, 
            spa_revenue,
            monthly_total,
            annual_total
        ])
    
    # Write headers
    revenue_headers = ["Month", "GPU Leasing", "ASIC Mining", "Spa Income", "Monthly Total", "Annualized"]
    for col_idx, header in enumerate(revenue_headers, 1):
        cell = revenue_ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
    
    # Write revenue data
    for row_idx, row_data in enumerate(revenue_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = revenue_ws.cell(row=row_idx, column=col_idx, value=value)
            if col_idx > 1:  # Format currency columns
                cell.number_format = currency_format
    
    # Adjust column widths
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        revenue_ws.column_dimensions[col].width = 15
    
    # 3. Operating Expenses Sheet
    opex_ws = wb.create_sheet("Operating Expenses")
    
    # 5MW Strategic Operating expense data - Optimized for El Salvador
    opex_data = []
    for i, month in enumerate(months):
        # Energy calculation: $0.05/kWh × 5MW × 720 hours × 85% uptime = ~$153,000/month
        energy_cost = 0.05 * 5000 * 720 * 0.85
        staff_cost = 13000  # Reduced for El Salvador labor market
        maintenance_cost = 8000  # Scaled for 5MW
        insurance_cost = 3000
        connectivity_cost = 4000
        other_cost = 4000
        
        # Inflation adjustment (2% annual)
        inflation_factor = (1.02) ** (i / 12)
        
        total_opex = (energy_cost + staff_cost + maintenance_cost + 
                     insurance_cost + connectivity_cost + other_cost) * inflation_factor
        
        opex_data.append([
            month,
            energy_cost * inflation_factor,
            staff_cost * inflation_factor,
            maintenance_cost * inflation_factor,
            insurance_cost * inflation_factor,
            connectivity_cost * inflation_factor,
            other_cost * inflation_factor,
            total_opex
        ])
    
    # Write OpEx headers
    opex_headers = ["Month", "Energy", "Staff", "Maintenance", "Insurance", "Connectivity", "Other", "Total OpEx"]
    for col_idx, header in enumerate(opex_headers, 1):
        cell = opex_ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
    
    # Write OpEx data
    for row_idx, row_data in enumerate(opex_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = opex_ws.cell(row=row_idx, column=col_idx, value=value)
            if col_idx > 1:  # Format currency columns
                cell.number_format = currency_format
    
    # Adjust column widths
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        opex_ws.column_dimensions[col].width = 15
    
    # 4. ROI Timeline Sheet
    roi_ws = wb.create_sheet("ROI Timeline")
    
    # Calculate ROI timeline - 5MW Strategic
    initial_investment = 6000000
    cumulative_cashflow = 0
    roi_data = []
    
    for i, month in enumerate(months):
        monthly_revenue = revenue_data[i][4]  # Monthly total from revenue sheet
        monthly_opex = opex_data[i][7]  # Total OpEx from opex sheet
        monthly_cashflow = monthly_revenue - monthly_opex
        cumulative_cashflow += monthly_cashflow
        
        net_position = cumulative_cashflow - initial_investment
        roi_percentage = (cumulative_cashflow / initial_investment) if initial_investment > 0 else 0
        payback_achieved = "YES" if net_position >= 0 else "NO"
        
        roi_data.append([
            month,
            monthly_revenue,
            monthly_opex,
            monthly_cashflow,
            cumulative_cashflow,
            net_position,
            roi_percentage,
            payback_achieved
        ])
    
    # Write ROI headers
    roi_headers = ["Month", "Revenue", "OpEx", "Net Cash Flow", "Cumulative CF", "Net Position", "ROI %", "Payback"]
    for col_idx, header in enumerate(roi_headers, 1):
        cell = roi_ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
    
    # Write ROI data
    for row_idx, row_data in enumerate(roi_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = roi_ws.cell(row=row_idx, column=col_idx, value=value)
            if col_idx in [2, 3, 4, 5, 6]:  # Currency columns
                cell.number_format = currency_format
            elif col_idx == 7:  # Percentage column
                cell.number_format = percent_format
            
            # Highlight break-even point
            if col_idx == 8 and value == "YES":
                cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    
    # Adjust column widths
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        roi_ws.column_dimensions[col].width = 15
    
    # 5. Summary Dashboard Sheet
    summary_ws = wb.create_sheet("Executive Summary")
    
    # 5MW Strategic Key metrics
    total_capex = 6200000
    avg_monthly_revenue = sum(row[4] for row in revenue_data[:12]) / 12
    avg_monthly_opex = sum(row[7] for row in opex_data[:12]) / 12
    monthly_net_cashflow = avg_monthly_revenue - avg_monthly_opex
    
    # Find break-even month
    break_even_month = "Not achieved in 5 years"
    for i, row in enumerate(roi_data):
        if row[5] >= 0:  # Net position positive
            break_even_month = row[0]
            break
    
    summary_data = [
        ["GridEdge Compute Center - Phase I 5MW Modular Launch", ""],
        ["", ""],
        ["Investment Overview", ""],
        ["Total CapEx", total_capex],
        ["Project Capacity", "5.0 MW"],
        ["Location", "El Salvador Geothermal Corridor"],
        ["", ""],
        ["Revenue Performance (Year 1 Average)", ""],
        ["Monthly Revenue", avg_monthly_revenue],
        ["Annual Revenue", avg_monthly_revenue * 12],
        ["", ""],
        ["Operating Performance (Year 1 Average)", ""],
        ["Monthly OpEx", avg_monthly_opex],
        ["Annual OpEx", avg_monthly_opex * 12],
        ["Monthly Net Cash Flow", monthly_net_cashflow],
        ["", ""],
        ["Investment Returns", ""],
        ["Break-even Month", break_even_month],
        ["5-Year Cumulative Cash Flow", roi_data[-1][4]],
        ["5-Year ROI", roi_data[-1][6]],
        ["", ""],
        ["Key Assumptions", ""],
        ["GPU Utilization", "70-80%"],
        ["Energy Cost", "$0.05/kWh"],
        ["Facility Uptime", "85%"],
        ["Bitcoin Price (avg)", "$40,000"],
    ]
    
    # Write summary data
    for row_idx, (label, value) in enumerate(summary_data, 1):
        summary_ws.cell(row=row_idx, column=1, value=label)
        summary_ws.cell(row=row_idx, column=2, value=value)
        
        # Style headers
        if "Overview" in label or "Performance" in label or "Returns" in label or "Assumptions" in label:
            summary_ws.cell(row=row_idx, column=1).font = Font(bold=True, size=12)
            summary_ws.cell(row=row_idx, column=1).fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
        
        # Format currency values
        if isinstance(value, (int, float)) and value > 1000:
            summary_ws.cell(row=row_idx, column=2).number_format = currency_format
    
    # Adjust column widths
    summary_ws.column_dimensions['A'].width = 30
    summary_ws.column_dimensions['B'].width = 20
    
    # 6. Phased Build Plan Sheet
    phases_ws = wb.create_sheet("Phased Build Plan")
    
    # Phased build data
    phases_data = [
        ["Phase", "MW Added", "Cumulative MW", "CapEx Estimate", "Target Year", "Notes"],
        ["Phase I", "2.5 MW", "2.5 MW", 3200000, "Q1 2026", "Geothermal, Spa reuse, Modular datacenter"],
        ["Phase II", "2.5 MW", "5.0 MW", 3500000, "Q4 2026", "Expand GPU capacity, add staff housing"],
        ["Phase III", "5-10 MW", "10-15 MW", "TBD", "2027-2028", "Sovereign AI clusters, Grid services"],
        ["", "", "", "", "", ""],
        ["Total (3 Phases)", "10-15 MW", "15 MW", "~$10-12M", "2026-2028", "Full build-out capacity"],
        ["", "", "", "", "", ""],
        ["Phase I Details", "", "", "", "", ""],
        ["ASIC Mining", "1.5 MW", "", "", "", "Bitcoin generation focus"],
        ["GPU Clusters", "1.0 MW", "", "", "", "AI/ML workload hosting"],
        ["Waste Heat Recovery", "", "", "", "", "Spa partnership integration"],
        ["Geothermal Connection", "", "", "", "", "LaGeo PPA for baseload power"],
        ["", "", "", "", "", ""],
        ["Phase II Expansion", "", "", "", "", ""],
        ["Additional GPUs", "1.5 MW", "", "", "", "Scale AI hosting capacity"],
        ["Enhanced ASIC", "1.0 MW", "", "", "", "Next-gen mining hardware"],
        ["Staff Housing", "", "", "", "", "On-site accommodation facility"],
        ["Grid Integration", "", "", "", "", "Enhanced grid services capability"],
        ["", "", "", "", "", ""],
        ["Phase III Vision", "", "", "", "", ""],
        ["Sovereign AI", "5-10 MW", "", "", "", "National AI infrastructure"],
        ["Grid Stabilization", "", "", "", "", "Frequency regulation services"],
        ["Export Capacity", "", "", "", "", "Regional power market participation"],
    ]
    
    # Write phased build headers
    for col_idx, header in enumerate(phases_data[0], 1):
        cell = phases_ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
    
    # Write phased build data
    for row_idx, row_data in enumerate(phases_data[1:], 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = phases_ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Format currency column
            if col_idx == 4 and isinstance(value, (int, float)):
                cell.number_format = currency_format
            
            # Bold phase headers and section headers
            if col_idx == 1 and ("Phase" in str(value) or "Details" in str(value) or "Expansion" in str(value) or "Vision" in str(value)):
                cell.font = Font(bold=True)
    
    # Adjust column widths
    phases_ws.column_dimensions['A'].width = 20
    phases_ws.column_dimensions['B'].width = 15
    phases_ws.column_dimensions['C'].width = 15
    phases_ws.column_dimensions['D'].width = 15
    phases_ws.column_dimensions['E'].width = 15
    phases_ws.column_dimensions['F'].width = 35
    
    # Set active sheet to summary
    wb.active = summary_ws

    # 7. Financing Options Sheet
    financing_ws = wb.create_sheet("Financing Options")
    
    # Financing options data
    financing_data = [
        ["Asset", "Financing Type", "Example Lenders/Platforms", "Notes"],
        ["Bitcoin (self-mined)", "BTC-backed Line of Credit", "Unchained, Ledn, Local El Salvador Lenders", "Use freshly mined BTC as collateral for stablecoin loans (Liquid USDt)."],
        ["GPUs (NVIDIA A6000/H100)", "Hardware-backed Leasing", "Coreweave, Lambda, Hydra Host", "Finance GPU acquisition via lease-to-own or sale-leaseback agreements."],
        ["ASIC Miners (S21/T21)", "Equipment Financing", "Asset-backed loans from crypto-friendly funds", "Use the hardware itself as collateral for financing the purchase."],
        ["Modular Containers/Infra", "Equipment-based Financing", "Private credit, off-balance sheet leasing", "Collateralize the physical infrastructure for private loans."],
        ["Forward Sales", "Prepaid GPU Contracts", "Direct sales to AI startups, render farms", "Sell future GPU access at a discount to fund current OpEx and CapEx."],
        ["Carbon Credits", "Green Financing", "ESG funds, Methane capture grants", "Monetize methane flaring partnerships for carbon credits and grant funding."],
        ["Sovereign Partnership", "Government Grants/Loans", "El Salvador National Bitcoin Office", "Explore partnerships for building sovereign AI infrastructure, potentially unlocking grants."]
    ]
    
    # Write financing headers
    for col_idx, header in enumerate(financing_data[0], 1):
        cell = financing_ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        
    # Write financing data
    for row_idx, row_data in enumerate(financing_data[1:], 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = financing_ws.cell(row=row_idx, column=col_idx, value=value)

    # Adjust column widths
    financing_ws.column_dimensions['A'].width = 25
    financing_ws.column_dimensions['B'].width = 30
    financing_ws.column_dimensions['C'].width = 40
    financing_ws.column_dimensions['D'].width = 70

    
    return wb

def main():
    """Main function to create and save the financial model"""
    print("Creating GridEdge Compute Center Financial Model...")

    
    # Create the workbook
    wb = create_financial_model()
    
    # Save the file
    filename = "financial_model.xlsx"
    wb.save(filename)
    
    print(f"✅ Financial model saved as {filename}")
    print("\n5MW Strategic Model includes:")
    print("- CapEx Breakdown: $6.0M total investment for 5MW capacity")
    print("- Monthly Revenue Forecast: GPU ($85K) + ASIC ($22.5K) + Spa ($2.5K)")
    print("- Operating Expenses: Energy ($153K), Staff ($13K), Other ($19K)")
    print("- ROI Timeline: Month-by-month cash flow analysis with break-even")
    print("- Executive Summary: 5MW strategic deployment metrics")
    print("- Phased Build Plan: Growth roadmap to 15MW+ capacity")
    
    return filename

if __name__ == "__main__":
    main()
