#!/usr/bin/env python3
"""
GridEdge Compute Center - ROI Timeline Chart Generator
Reads from financial_model.xlsx and creates break-even visualization
"""

import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from openpyxl import load_workbook
import numpy as np
import os
from datetime import datetime, timedelta

def read_roi_data(excel_file):
    """Read ROI timeline data from the Excel file"""
    
    # Load the workbook and get the ROI Timeline sheet
    wb = load_workbook(excel_file)
    ws = wb["ROI Timeline"]
    
    # Extract data from the worksheet
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):  # Skip header
        if row[0] and row[4]:  # Check if we have month and cumulative cash flow
            data.append({
                'month': row[0],
                'monthly_revenue': row[1] if row[1] else 0,
                'monthly_opex': row[2] if row[2] else 0,
                'net_cashflow': row[3] if row[3] else 0,
                'cumulative_cf': row[4] if row[4] else 0,
                'net_position': row[5] if row[5] else 0,
                'roi_percent': row[6] if row[6] else 0,
                'payback': row[7] if row[7] else "NO"
            })
    
    df = pd.DataFrame(data)
    return df

def create_roi_chart(df, output_path):
    """Create ROI timeline chart showing break-even forecast"""
    
    # Create figure with larger size for better readability
    fig, ax = plt.subplots(figsize=(14, 8))
    
    # Convert month strings to numeric values for plotting
    month_numbers = range(1, len(df) + 1)
    
    # Plot cumulative cash flow
    line = ax.plot(month_numbers, df['cumulative_cf'] / 1000000, 
                   linewidth=3, color='#1f4e79', label='Cumulative Cash Flow')
    
    # Add horizontal line at y=0 (break-even)
    ax.axhline(y=0, color='red', linestyle='--', linewidth=2, alpha=0.7, label='Break-even Line')
    
    # Find break-even point or maximum loss
    net_positions = df['net_position'].values
    break_even_month = None
    for i, net_pos in enumerate(net_positions):
        if net_pos >= 0:
            break_even_month = i + 1
            break
    
    # Highlight break-even point or point of maximum loss
    if break_even_month:
        # Mark break-even point
        break_even_cf = df.iloc[break_even_month - 1]['cumulative_cf'] / 1000000
        ax.plot(break_even_month, break_even_cf, 'go', markersize=12, 
                label=f'Break-even: Month {break_even_month}')
        
        # Add annotation
        ax.annotate(f'Break-even\nMonth {break_even_month}', 
                   xy=(break_even_month, break_even_cf),
                   xytext=(break_even_month + 6, break_even_cf + 0.5),
                   arrowprops=dict(arrowstyle='->', color='green', lw=2),
                   fontsize=11, fontweight='bold', color='green',
                   bbox=dict(boxstyle="round,pad=0.3", facecolor='lightgreen', alpha=0.7))
    else:
        # Mark maximum loss point (lowest cumulative CF)
        min_cf_idx = df['cumulative_cf'].idxmin()
        min_cf_month = min_cf_idx + 1
        min_cf_value = df.iloc[min_cf_idx]['cumulative_cf'] / 1000000
        
        ax.plot(min_cf_month, min_cf_value, 'ro', markersize=12, 
                label=f'Maximum Loss: Month {min_cf_month}')
        
        # Add annotation
        ax.annotate(f'Maximum Loss\nMonth {min_cf_month}\n${min_cf_value:.1f}M', 
                   xy=(min_cf_month, min_cf_value),
                   xytext=(min_cf_month + 6, min_cf_value - 0.5),
                   arrowprops=dict(arrowstyle='->', color='red', lw=2),
                   fontsize=11, fontweight='bold', color='red',
                   bbox=dict(boxstyle="round,pad=0.3", facecolor='lightcoral', alpha=0.7))
    
    # Customize the chart
    ax.set_xlabel('Month', fontsize=12, fontweight='bold')
    ax.set_ylabel('Cumulative Cash Flow (Millions USD)', fontsize=12, fontweight='bold')
    ax.set_title('GridEdge Phase I – Break-even Forecast\n2.5MW Modular Data Center', 
                 fontsize=16, fontweight='bold', pad=20)
    
    # Add grid for easier reading
    ax.grid(True, alpha=0.3, linestyle='--')
    
    # Format x-axis
    ax.set_xlim(0, len(df) + 2)
    year_ticks = list(range(0, len(df) + 1, 12))
    ax.set_xticks(year_ticks)  # Show every 12 months
    ax.set_xticklabels([f'Year {i}' if i > 0 else 'Start' for i in range(len(year_ticks))])
    
    # Add minor ticks for months
    ax.set_xticks(range(1, len(df) + 1), minor=True)
    
    # Format y-axis to show millions
    ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: f'${x:.1f}M'))
    
    # Add legend
    ax.legend(loc='upper left', fontsize=10, framealpha=0.9)
    
    # Add summary text box
    final_cf = df.iloc[-1]['cumulative_cf'] / 1000000
    final_roi = df.iloc[-1]['roi_percent'] * 100
    avg_monthly_net = df['net_cashflow'].mean()
    
    summary_text = f"""5-Year Summary:
• Final Cash Flow: ${final_cf:.1f}M
• ROI: {final_roi:.1f}%
• Avg Monthly Net: ${avg_monthly_net/1000:.0f}K
• CapEx: $3.2M"""
    
    ax.text(0.02, 0.98, summary_text, transform=ax.transAxes, 
            verticalalignment='top', fontsize=10,
            bbox=dict(boxstyle="round,pad=0.4", facecolor='lightblue', alpha=0.8))
    
    # Adjust layout
    plt.tight_layout()
    
    # Save the chart
    plt.savefig(output_path, dpi=300, bbox_inches='tight', facecolor='white')
    print(f"✅ ROI timeline chart saved to: {output_path}")
    
    return fig, break_even_month

def create_detailed_cashflow_chart(df, output_path):
    """Create detailed monthly cash flow chart showing revenue vs expenses"""
    
    fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(14, 10))
    
    month_numbers = range(1, len(df) + 1)
    
    # Top chart: Monthly Revenue vs OpEx
    ax1.bar(month_numbers, df['monthly_revenue'] / 1000, 
            alpha=0.7, color='green', label='Monthly Revenue', width=0.8)
    ax1.bar(month_numbers, -df['monthly_opex'] / 1000, 
            alpha=0.7, color='red', label='Monthly OpEx', width=0.8)
    
    ax1.set_ylabel('Monthly Cash Flow ($K)', fontsize=11, fontweight='bold')
    ax1.set_title('GridEdge Phase I – Monthly Revenue vs Operating Expenses', 
                  fontsize=14, fontweight='bold')
    ax1.grid(True, alpha=0.3, linestyle='--')
    ax1.legend(loc='upper right')
    ax1.axhline(y=0, color='black', linewidth=1)
    
    # Bottom chart: Net Cash Flow
    colors = ['red' if x < 0 else 'green' for x in df['net_cashflow']]
    ax2.bar(month_numbers, df['net_cashflow'] / 1000, 
            color=colors, alpha=0.7, width=0.8)
    
    ax2.set_xlabel('Month', fontsize=11, fontweight='bold')
    ax2.set_ylabel('Net Cash Flow ($K)', fontsize=11, fontweight='bold')
    ax2.set_title('Monthly Net Cash Flow (Revenue - OpEx)', fontsize=14, fontweight='bold')
    ax2.grid(True, alpha=0.3, linestyle='--')
    ax2.axhline(y=0, color='black', linewidth=1)
    
    # Format x-axis for both charts
    for ax in [ax1, ax2]:
        ax.set_xlim(0, len(df) + 1)
        year_ticks = list(range(0, len(df) + 1, 12))
        ax.set_xticks(year_ticks)
        ax.set_xticklabels([f'Y{i}' if i > 0 else 'Start' for i in range(len(year_ticks))])
    
    plt.tight_layout()
    plt.savefig(output_path, dpi=300, bbox_inches='tight', facecolor='white')
    print(f"✅ Detailed cash flow chart saved to: {output_path}")
    
    return fig

def main():
    """Main function to generate ROI charts"""
    
    # File paths
    excel_file = "financial_model.xlsx"
    graphs_dir = "graphs"
    roi_chart_path = os.path.join(graphs_dir, "roi_chart.png")
    detailed_chart_path = os.path.join(graphs_dir, "cashflow_detailed.png")
    
    # Check if Excel file exists
    if not os.path.exists(excel_file):
        print(f"❌ Error: {excel_file} not found. Please run create_financial_model.py first.")
        return
    
    # Create graphs directory if it doesn't exist
    os.makedirs(graphs_dir, exist_ok=True)
    
    print("📈 Generating ROI and cash flow charts...")
    
    try:
        # Read data from Excel
        df = read_roi_data(excel_file)
        print(f"📊 Loaded {len(df)} months of ROI data from Excel")
        
        # Generate main ROI chart
        fig1, break_even_month = create_roi_chart(df, roi_chart_path)
        
        # Generate detailed cash flow chart
        fig2 = create_detailed_cashflow_chart(df, detailed_chart_path)
        
        print("\n🎯 Chart generation complete!")
        print(f"   • ROI timeline: {roi_chart_path}")
        print(f"   • Cash flow detail: {detailed_chart_path}")
        
        # Display key insights
        print("\n💡 Key Insights:")
        if break_even_month:
            print(f"   • Break-even achieved: Month {break_even_month}")
            years = break_even_month // 12
            months = break_even_month % 12
            if years > 0:
                print(f"   • Payback period: {years} year(s) {months} month(s)")
            else:
                print(f"   • Payback period: {months} month(s)")
        else:
            print("   • Break-even not achieved in 5-year model")
        
        final_cf = df.iloc[-1]['cumulative_cf']
        final_roi = df.iloc[-1]['roi_percent'] * 100
        avg_monthly_net = df['net_cashflow'].mean()
        
        print(f"   • 5-year cumulative cash flow: ${final_cf/1000000:.1f}M")
        print(f"   • 5-year ROI: {final_roi:.1f}%")
        print(f"   • Average monthly net cash flow: ${avg_monthly_net/1000:.0f}K")
        
        # Monthly cash flow trend
        positive_months = len(df[df['net_cashflow'] > 0])
        print(f"   • Months with positive cash flow: {positive_months}/{len(df)} ({positive_months/len(df)*100:.1f}%)")
            
    except Exception as e:
        print(f"❌ Error generating charts: {str(e)}")
        raise

if __name__ == "__main__":
    main()
