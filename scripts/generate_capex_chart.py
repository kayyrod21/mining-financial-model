#!/usr/bin/env python3
"""
GridEdge Compute Center - CapEx Breakdown Chart Generator
Reads from financial_model.xlsx and creates visual charts for the CapEx breakdown
"""

import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from openpyxl import load_workbook
import numpy as np
import os

def read_capex_data(excel_file):
    """Read CapEx data from the Excel file"""
    
    # Load the workbook and get the CapEx sheet
    wb = load_workbook(excel_file)
    ws = wb["CapEx Breakdown"]
    
    # Extract data from the worksheet
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):  # Skip header
        if row[0] and row[1] and row[4]:  # Check if we have category, subcategory, and total cost
            if isinstance(row[4], (int, float)) and row[4] > 0:
                data.append({
                    'category': row[0],
                    'subcategory': row[1], 
                    'total_cost': row[4]
                })
    
    df = pd.DataFrame(data)
    return df

def create_capex_pie_chart(df, output_path):
    """Create a pie chart showing CapEx breakdown by major categories"""
    
    # Group by major categories
    category_totals = df.groupby('category')['total_cost'].sum()
    
    # Filter out zero or negative values
    category_totals = category_totals[category_totals > 0]
    
    # Calculate total for percentage labels
    total_capex = category_totals.sum()
    
    # Create figure and axis
    fig, ax = plt.subplots(figsize=(12, 8))
    
    # Color scheme - professional blue/gray palette
    colors = ['#1f4e79', '#2e75b6', '#4a90c2', '#7bb3d0', '#a6d0e4', '#d4e8f0']
    
    # Create pie chart
    wedges, texts, autotexts = ax.pie(
        category_totals.values,
        labels=category_totals.index,
        colors=colors[:len(category_totals)],
        autopct=lambda pct: f'${category_totals.values[int(pct/100*len(category_totals))]//1000000:.1f}M\n({pct:.1f}%)',
        startangle=90,
        textprops={'fontsize': 10, 'weight': 'bold'}
    )
    
    # Enhance text styling
    for autotext in autotexts:
        autotext.set_color('white')
        autotext.set_weight('bold')
    
    # Add title
    plt.title('GridEdge Compute Center - CapEx Breakdown by Category\nTotal Investment: ${:.1f}M'.format(total_capex/1000000), 
              fontsize=16, weight='bold', pad=20)
    
    # Add legend with detailed breakdown
    legend_labels = []
    for category, total in category_totals.items():
        percentage = (total / total_capex) * 100
        legend_labels.append(f'{category}: ${total/1000000:.1f}M ({percentage:.1f}%)')
    
    plt.legend(wedges, legend_labels, title="Investment Categories", 
               loc="center left", bbox_to_anchor=(1, 0, 0.5, 1), fontsize=10)
    
    # Ensure equal aspect ratio for circular pie
    ax.axis('equal')
    
    # Adjust layout to prevent legend cutoff
    plt.tight_layout()
    
    # Save the chart
    plt.savefig(output_path, dpi=300, bbox_inches='tight', facecolor='white')
    print(f"✅ CapEx pie chart saved to: {output_path}")
    
    return fig

def create_capex_bar_chart(df, output_path):
    """Create a horizontal bar chart showing detailed CapEx breakdown"""
    
    # Sort by total cost for better visualization
    df_sorted = df.sort_values('total_cost', ascending=True)
    
    # Create figure
    fig, ax = plt.subplots(figsize=(12, 10))
    
    # Color mapping for categories
    category_colors = {
        'Equipment': '#1f4e79',
        'Facility': '#2e75b6', 
        'Power & Cooling': '#4a90c2',
        'Contingency': '#7bb3d0'
    }
    
    # Create colors list based on categories
    colors = [category_colors.get(cat, '#a6d0e4') for cat in df_sorted['category']]
    
    # Create horizontal bar chart
    bars = ax.barh(range(len(df_sorted)), df_sorted['total_cost'] / 1000000, color=colors)
    
    # Customize y-axis labels
    ax.set_yticks(range(len(df_sorted)))
    ax.set_yticklabels([f"{row['category']}\n{row['subcategory']}" for _, row in df_sorted.iterrows()], 
                       fontsize=9)
    
    # Add value labels on bars
    for i, (bar, cost) in enumerate(zip(bars, df_sorted['total_cost'])):
        width = bar.get_width()
        ax.text(width + 0.1, bar.get_y() + bar.get_height()/2, 
                f'${cost/1000000:.1f}M', 
                ha='left', va='center', fontsize=9, weight='bold')
    
    # Customize chart
    ax.set_xlabel('Investment Amount (Millions USD)', fontsize=12, weight='bold')
    ax.set_title('GridEdge Compute Center - Detailed CapEx Breakdown', 
                 fontsize=14, weight='bold', pad=20)
    
    # Add grid for easier reading
    ax.grid(axis='x', alpha=0.3, linestyle='--')
    
    # Create custom legend
    legend_elements = [mpatches.Patch(color=color, label=category) 
                      for category, color in category_colors.items()]
    ax.legend(handles=legend_elements, loc='lower right', fontsize=10)
    
    # Adjust layout
    plt.tight_layout()
    
    # Save the chart
    plt.savefig(output_path, dpi=300, bbox_inches='tight', facecolor='white')
    print(f"✅ CapEx bar chart saved to: {output_path}")
    
    return fig

def main():
    """Main function to generate CapEx charts"""
    
    # File paths
    excel_file = "financial_model.xlsx"
    graphs_dir = "graphs"
    pie_chart_path = os.path.join(graphs_dir, "capex_breakdown.png")
    bar_chart_path = os.path.join(graphs_dir, "capex_detailed.png")
    
    # Check if Excel file exists
    if not os.path.exists(excel_file):
        print(f"❌ Error: {excel_file} not found. Please run create_financial_model.py first.")
        return
    
    # Create graphs directory if it doesn't exist
    os.makedirs(graphs_dir, exist_ok=True)
    
    print("📊 Generating CapEx breakdown charts...")
    
    try:
        # Read data from Excel
        df = read_capex_data(excel_file)
        print(f"📈 Loaded {len(df)} CapEx items from Excel")
        
        # Generate pie chart
        create_capex_pie_chart(df, pie_chart_path)
        
        # Generate detailed bar chart
        create_capex_bar_chart(df, bar_chart_path)
        
        print("\n🎯 Chart generation complete!")
        print(f"   • Pie chart: {pie_chart_path}")
        print(f"   • Bar chart: {bar_chart_path}")
        
        # Display summary
        total_capex = df['total_cost'].sum()
        print(f"\n💰 Total CapEx: ${total_capex/1000000:.1f}M")
        print("📋 Category breakdown:")
        for category, total in df.groupby('category')['total_cost'].sum().items():
            percentage = (total / total_capex) * 100
            print(f"   • {category}: ${total/1000000:.1f}M ({percentage:.1f}%)")
            
    except Exception as e:
        print(f"❌ Error generating charts: {str(e)}")
        raise

if __name__ == "__main__":
    main()
