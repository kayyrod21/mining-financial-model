#!/usr/bin/env python3
"""
GridEdge Compute Center - Scenario Modeling
Generates scenario-based financial projections
"""

import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook, Workbook
import os

def read_roi_data(excel_file, sheet_name="ROI Timeline"):
    """Read ROI timeline data from the Excel file"""
    
    # Load the workbook and get the ROI Timeline sheet
    wb = load_workbook(excel_file)
    ws = wb[sheet_name]
    
    # Extract data from the worksheet
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and row[4]:  # Check if we have month and cumulative cash flow
            data.append({
                'month': row[0],
                'monthly_revenue': row[1] if row[1] else 0,
                'monthly_opex': row[2] if row[2] else 0,
                'net_cashflow': row[3] if row[3] else 0,
                'cumulative_cf': row[4] if row[4] else 0,
                'net_position': row[5] if row[5] else 0,
                'roi_percent': row[6] if row[6] else 0,
                'payback': row[7] if row[7] else "NO"
            })
    
    df = pd.DataFrame(data)
    return df

def create_roi_chart(df, output_path, chart_title):
    """Create ROI timeline chart showing break-even forecast"""
    
    fig, ax = plt.subplots(figsize=(14, 8))
    
    month_numbers = range(1, len(df) + 1)
    
    # Plot cumulative cash flow
    ax.plot(month_numbers, df['cumulative_cf'] / 1000000, 
            linewidth=3, color='#1f4e79', label='Cumulative Cash Flow')
    
    # Add horizontal line at y=0 (break-even)
    ax.axhline(y=0, color='red', linestyle='--', linewidth=2, alpha=0.7, label='Break-even Line')
    
    # Find break-even point or maximum loss
    net_positions = df['net_position'].values
    break_even_month = None
    for i, net_pos in enumerate(net_positions):
        if net_pos >= 0:
            break_even_month = i + 1
            break
    
    # Highlight break-even point or point of maximum loss
    if break_even_month:
        # Mark break-even point
        break_even_cf = df.iloc[break_even_month - 1]['cumulative_cf'] / 1000000
        ax.plot(break_even_month, break_even_cf, 'go', markersize=12, 
                label=f'Break-even: Month {break_even_month}')
        
        # Add annotation
        ax.annotate(f'Break-even
Month {break_even_month}', 
                   xy=(break_even_month, break_even_cf),
                   xytext=(break_even_month + 1, break_even_cf + 0.5),
                   arrowprops=dict(arrowstyle='->', color='green', lw=2),
                   fontsize=11, fontweight='bold', color='green',
                   bbox=dict(boxstyle="round,pad=0.3", facecolor='lightgreen', alpha=0.7))
    else:
        # Mark maximum loss point (lowest cumulative CF)
        min_cf_idx = df['cumulative_cf'].idxmin()
        min_cf_month = min_cf_idx + 1
        min_cf_value = df.iloc[min_cf_idx]['cumulative_cf'] / 1000000
        
        ax.plot(min_cf_month, min_cf_value, 'ro', markersize=12, 
                label=f'Maximum Loss: Month {min_cf_month}')
        
        # Add annotation
        ax.annotate(f'Maximum Loss
Month {min_cf_month}
${min_cf_value:.1f}M', 
                   xy=(min_cf_month, min_cf_value),
                   xytext=(min_cf_month + 6, min_cf_value - 0.5),
                   arrowprops=dict(arrowstyle='->', color='red', lw=2),
                   fontsize=11, fontweight='bold', color='red',
                   bbox=dict(boxstyle="round,pad=0.3", facecolor='lightcoral', alpha=0.7))
    
    # Customize the chart
    ax.set_xlabel('Month', fontsize=12, fontweight='bold')
    ax.set_ylabel('Cumulative Cash Flow (Millions USD)', fontsize=12, fontweight='bold')
    ax.set_title(chart_title, 
                 fontsize=16, fontweight='bold', pad=20)
    
    # Add grid for easier reading
    ax.grid(True, alpha=0.3, linestyle='--')
    
    # Format x-axis
    ax.set_xlim(0, len(df) + 2)
    year_ticks = list(range(0, len(df) + 1, 12))
    ax.set_xticks(year_ticks)
    ax.set_xticklabels([f'Year {i}' if i > 0 else 'Start' for i in range(len(year_ticks))])
    
    # Add minor ticks for months
    ax.set_xticks(range(1, len(df) + 1), minor=True)
    
    # Format y-axis to show millions
    ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: f'${x:.1f}M'))
    
    # Add legend
    ax.legend(loc='upper left', fontsize=10, framealpha=0.9)
    
    # Add summary text box
    final_cf = df.iloc[-1]['cumulative_cf'] / 1000000
    final_roi = df.iloc[-1]['roi_percent'] * 100
    avg_monthly_net = df['net_cashflow'].mean()
    
    summary_text = f"""5-Year Summary:
• Final Cash Flow: ${final_cf:.1f}M
• ROI: {final_roi:.1f}%
• Avg Monthly Net: ${avg_monthly_net/1000:.0f}K
• CapEx: $6.2M"""
    
    ax.text(0.02, 0.98, summary_text, transform=ax.transAxes, 
            verticalalignment='top', fontsize=10,
            bbox=dict(boxstyle="round,pad=0.4", facecolor='lightblue', alpha=0.8))
    
    # Adjust layout
    plt.tight_layout()
    
    # Save the chart
    plt.savefig(output_path, dpi=300, bbox_inches='tight', facecolor='white')
    print(f"✅ Scenario chart saved to: {output_path}")
    
    return fig, break_even_month

def create_detailed_cashflow_chart(df, output_path):
    """Create detailed monthly cash flow chart showing revenue vs expenses"""
    
    fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(14, 10))
    
    month_numbers = range(1, len(df) + 1)
    
    # Top chart: Monthly Revenue vs OpEx
    ax1.bar(month_numbers, df['monthly_revenue'] / 1000, 
            alpha=0.7, color='green', label='Monthly Revenue', width=0.8)
    ax1.bar(month_numbers, -df['monthly_opex'] / 1000, 
            alpha=0.7, color='red', label='Monthly OpEx', width=0.8)
    
    ax1.set_ylabel('Monthly Cash Flow ($K)', fontsize=11, fontweight='bold')
    ax1.set_title('GridEdge Phase I – Monthly Revenue vs Operating Expenses', 
                  fontsize=14, fontweight='bold')
    ax1.grid(True, alpha=0.3, linestyle='--')
    ax1.legend(loc='upper right')
    ax1.axhline(y=0, color='black', linewidth=1)
    
    # Bottom chart: Net Cash Flow
    colors = ['red' if x < 0 else 'green' for x in df['net_cashflow']]
    ax2.bar(month_numbers, df['net_cashflow'] / 1000, 
            color=colors, alpha=0.7, width=0.8)
    
    ax2.set_xlabel('Month', fontsize=11, fontweight='bold')
    ax2.set_ylabel('Net Cash Flow ($K)', fontsize=11, fontweight='bold')
    ax2.set_title('Monthly Net Cash Flow (Revenue - OpEx)', fontsize=14, fontweight='bold')
    ax2.grid(True, alpha=0.3, linestyle='--')
    ax2.axhline(y=0, color='black', linewidth=1)
    
    # Format x-axis for both charts
    for ax in [ax1, ax2]:
        ax.set_xlim(0, len(df) + 1)
        ax.set_xticks(range(0, len(df) + 1, 12))
        ax.set_xticklabels([f'Y{i}' if i > 0 else 'Start' for i in range(0, len(df) // 12 + 2)])
    
    plt.tight_layout()
    plt.savefig(output_path, dpi=300, bbox_inches='tight', facecolor='white')
    print(f"✅ Detailed cash flow chart saved to: {output_path}")
    
    return fig

def generate_scenario_models(excel_file="financial_model.xlsx", graphs_dir="graphs"):
    """Generate and save charts for different scenarios."""
    
    # Ensure graphs directory exists
    os.makedirs(graphs_dir, exist_ok=True)
    
    # Load the base model data
    try:
      base_df = read_roi_data(excel_file, "ROI Timeline")
    except FileNotFoundError:
        print(f"❌ Error: {excel_file} not found. Please run create_financial_model.py first.")
        return
    
    if base_df.empty:
      print(f"❌ Error: ROI Timeline sheet is empty.")
      return

    # --- Bear Case ---
    bear_df = base_df.copy()
    bear_df['monthly_revenue'] = 80000 + bear_df['asic_revenue']
    bear_df['monthly_opex'] = 153000
    bear_df['net_cashflow'] = bear_df['monthly_revenue'] - bear_df['monthly_opex']
    bear_df['cumulative_cf'] = bear_df['net_cashflow'].cumsum()
    bear_df['net_position'] = bear_df['cumulative_cf'] - 6200000
    bear_df['roi_percent'] = (bear_df['cumulative_cf'] / 6200000) * 100
    bear_df['payback'] = bear_df['net_position'].apply(lambda x: "YES" if x >= 0 else "NO")
    
    # Calculate Bear Break-even
    net_positions = bear_df['net_position'].values
    bear_break_even_month = None
    for i, net_pos in enumerate(net_positions):
        if net_pos >= 0:
            bear_break_even_month = i + 1
            break

    bear_chart_path = os.path.join(graphs_dir, "scenario_bear.png")
    create_roi_chart(bear_df, bear_chart_path, chart_title="GridEdge 5MW – Bear Case (Low Utilization, High Power Costs)")
    
    # --- Base Case ---
    base_df = base_df.copy()
    base_df['monthly_revenue'] = 120000 + base_df['asic_revenue']
    base_df['monthly_opex'] = 138000  # Reduced due to uptime
    base_df['net_cashflow'] = base_df['monthly_revenue'] - base_df['monthly_opex']
    base_df['cumulative_cf'] = base_df['net_cashflow'].cumsum()
    base_df['net_position'] = base_df['cumulative_cf'] - 6200000
    base_df['roi_percent'] = (base_df['cumulative_cf'] / 6200000) * 100
    base_df['payback'] = base_df['net_position'].apply(lambda x: "YES" if x >= 0 else "NO")
    
    # Calculate Base Break-even
    net_positions = base_df['net_position'].values
    base_break_even_month = None
    for i, net_pos in enumerate(net_positions):
        if net_pos >= 0:
            base_break_even_month = i + 1
            break
    
    base_chart_path = os.path.join(graphs_dir, "scenario_base.png")
    create_roi_chart(base_df, base_chart_path, chart_title="GridEdge 5MW – Base Case")

    # --- Bull Case ---
    bull_df = base_df.copy()
    bull_df['monthly_revenue'] = 250000 + bull_df['asic_revenue']
    bull_df['monthly_opex'] = 116000
    bull_df['net_cashflow'] = bull_df['monthly_revenue'] - bull_df['monthly_opex']
    bull_df['cumulative_cf'] = bull_df['net_cashflow'].cumsum()
    bull_df['net_position'] = bull_df['cumulative_cf'] - 6200000
    bull_df['roi_percent'] = (bull_df['cumulative_cf'] / 6200000) * 100
    bull_df['payback'] = bull_df['net_position'].apply(lambda x: "YES" if x >= 0 else "NO")
    
    # Calculate Bull Break-even
    net_positions = bull_df['net_position'].values
    bull_break_even_month = None
    for i, net_pos in enumerate(net_positions):
        if net_pos >= 0:
            bull_break_even_month = i + 1
            break
    
    bull_chart_path = os.path.join(graphs_dir, "scenario_bull.png")
    create_roi_chart(bull_df, bull_chart_path, chart_title="GridEdge 5MW – Bull Case (High Utilization, Low Power Costs)")

    print("
Scenario Charts Generated:")
    print(f"   • Bear Case: {bear_chart_path}")
    print(f"   • Base Case: {base_chart_path}")
    print(f"   • Bull Case: {bull_chart_path}")

    return


def main():
    """Main function to generate scenario charts"""
    generate_scenario_models()

if __name__ == "__main__":
    main()
